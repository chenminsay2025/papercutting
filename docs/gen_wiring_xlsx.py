"""Generate docs/wiring-table.xlsx from wiring-table.md content."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent / "wiring-table.xlsx"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2563EB")
cat_fill = PatternFill("solid", fgColor="EFF6FF")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(style="thin", color="CBD5E1")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_header(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border


def style_row(ws, row: int, cols: int, alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor="F8FAFC") if alt else None
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.alignment = wrap
        cell.border = border
        if fill:
            cell.fill = fill


def set_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


ROWS = [
    ("一、通信与烧录", "01", "USB-TTL TX", "STM32 PA10（USART1_RX）", "电脑发、MCU 收", "与 RX 交叉接线；模块 3.3V TTL；115200 8N1"),
    ("", "02", "USB-TTL RX", "STM32 PA9（USART1_TX）", "MCU 发、电脑收", "与 TX 交叉接线；VCC 勿接 MCU"),
    ("", "03", "USB-TTL GND", "STM32 GND", "串口共地", "必接；与 24V 负极、继电器 DC− 共地"),
    ("", "04", "USB-TTL VCC", "（不接 MCU）", "模块自供电", "拨 3.3V 档；勿给 STM32 反向供电"),
    ("", "05", "ST-LINK SWDIO", "STM32 PA13", "SWD 烧录", "可与 USB-TTL 同时插；线色建议绿"),
    ("", "06", "ST-LINK SWCLK", "STM32 PA14", "SWD 烧录", "BOOT0 接 GND；线色建议蓝"),
    ("", "07", "ST-LINK GND", "STM32 GND", "烧录共地", "必接"),
    ("", "08", "ST-LINK 3.3V", "STM32 3.3V", "烧录供电", "可选；板子已有 3.3V 时可不接"),
    ("二、电源与共地", "09", "220V AC", "24V 开关电源", "主电源", "电流 ≥ 电机额定，建议留 30% 余量"),
    ("", "10", "24V+", "NO1 与 NO2（短接拱桥）", "电机 H 桥正极", "24V 只接触点侧；勿接继电器 DC+"),
    ("", "11", "24V−（GND）", "NC1 与 NC2（短接拱桥）", "电机 H 桥负极", "与 STM32、TTL、ST-LINK、继电器 DC− 共地"),
    ("", "12", "24V−", "系统 GND 总线", "共地参考", "所有 GND 汇一点；24V+ 禁止接 MCU 引脚"),
    ("", "13", "12V 电源+", "继电器 DC+", "继电器线圈/光耦", "12V 模块接 12V，5V 模块接 5V；勿接 24V"),
    ("", "14", "系统 GND", "继电器 DC−", "继电器控制回路地", "与 24V−、STM32 GND 相连"),
    ("", "15", "3.3V 电源", "STM32 3.3V", "MCU 供电", "禁止 24V 直连；可与 ST-LINK 3.3V 二选一"),
    ("三、继电器控制线", "16", "跳线 S1~S4", "插 H 位", "高电平触发", "与当前固件一致；插 L 须改固件或改回 H"),
    ("", "17", "STM32 PA0", "继电器 IN1（K1）", "伸缩杆缩回", "缩回时仅 K1 吸合（PA0=1，PA1=0）"),
    ("", "18", "STM32 PA1", "继电器 IN2（K2）", "伸缩杆伸出", "伸出时仅 K2 吸合（PA0=0，PA1=1）；换向前先停 80ms"),
    ("", "19", "STM32 PA2", "继电器 IN3（K3）", "脉冲「继续」", "软件 PULSE_A；默认约 200ms"),
    ("", "20", "STM32 PA3", "继电器 IN4（K4）", "脉冲「原点」", "软件 PULSE_B；默认约 200ms"),
    ("四、伸缩杆电机", "21", "NO1", "NO2（短接）", "24V+ 分配到 K1/K2", "端子上用短导线连在一起"),
    ("", "22", "NC1", "NC2（短接）", "GND 分配到 K1/K2", "与系统 GND、24V− 共地"),
    ("", "23", "伸缩杆线 A", "COM1", "电机端 A", "负载接 COM；勿把 NC/NO 跨并到同一电机线"),
    ("", "24", "伸缩杆线 B", "COM2", "电机端 B", "方向反了则对调 COM1/COM2 两根线"),
    ("", "25", "NC2", "（悬空）", "未使用", "不接；旧方案误接会导致停止仍通电"),
    ("", "26", "—", "—", "停止状态 PA0=0 PA1=0", "K1/K2 释放；COM1、COM2 经 NC 接 GND，电机无电压"),
    ("五、切纸机按钮", "27", "切纸机「继续」按钮一端", "COM3", "并联继续键", "须低压干触点；220V 须光耦隔离"),
    ("", "28", "切纸机「继续」按钮另一端", "NO3", "并联继续键", "K3 吸合时 COM3–NO3 闭合，模拟按一下"),
    ("", "29", "NC3", "（悬空）", "未使用", "不接"),
    ("", "30", "切纸机「原点」按钮一端", "COM4", "并联原点键", "万用表确认干触点后再并"),
    ("", "31", "切纸机「原点」按钮另一端", "NO4", "并联原点键", "K4 吸合时 COM4–NO4 闭合，模拟按一下"),
    ("", "32", "NC4", "（悬空）", "未使用", "不接"),
    ("六、按键与指示灯", "33", "STM32 PB8", "切换按钮 → GND", "现场手动伸缩", "nologo 一体板；PA7=LCD RES；按一次缩回 3s，再按伸出 3s"),
    ("", "34", "（nologo）PA6", "LCD 背光 BLK", "板载屏背光", "低电平亮；勿接 D3 串口 LED"),
    ("", "35", "STM32 PA4", "D1 缩回 LED → 330Ω → GND", "缩回指示", "缩回时常亮"),
    ("", "36", "STM32 PA5", "D2 伸出 LED → 330Ω → GND", "伸出指示", "伸出时常亮"),
    ("", "37", "STM32 PB9", "D3 串口 LED → 330Ω → GND", "通信状态", "nologo 板；标准板仍用 PA6；快闪=未连接，慢呼吸=已连接"),
    ("六点五、槽型光电", "45", "槽型光电 VCC", "STM32 3.3V", "传感器供电", "模块 3.3~5V"),
    ("", "46", "槽型光电 GND", "系统 GND", "传感器共地", "与 STM32、继电器共地"),
    ("", "47", "槽型光电 DO", "STM32 PA8", "压纸检测", "遮挡=压纸中/ROD:HOME；未遮挡=未压纸/ROD:AWAY；缩回中自动停电机"),
    ("", "48", "槽型光电 AO", "（悬空）", "未使用", "本模块 AO 不起作用"),
    ("六点六、板载 LCD", "49", "PB0", "LCD DC", "数据/命令", "nologo 一体板已焊；勿占用"),
    ("", "50", "PB1", "LCD CS", "片选", "nologo 一体板已焊"),
    ("", "51", "PB10", "LCD SCL", "时钟", "nologo 一体板已焊"),
    ("", "52", "PB11", "LCD SDA/MOSI", "数据", "nologo 一体板已焊"),
    ("", "53", "PA7", "LCD RES", "复位", "nologo 一体板已焊；勿接切换键"),
    ("", "54", "PA6", "LCD 背光 BLK", "背光", "低电平亮；与 D3 串口灯不可共用"),
    ("", "55", "PB12~PB15", "W25Q64 Flash SPI2", "板载 Flash", "固件勿改这些脚"),
    ("七、软件", "38", "CutPPaper", "Cutting Master 4", "发送切割", "默认 Ctrl+P；须先激活目标窗口"),
    ("", "39", "CutPPaper", "USB-TTL COM 口", "流程控制", "设备管理器确认 COM 号；烧录/拔 USB 后须重新连接"),
    ("", "40", "CutPPaper", "—", "与 MCU 通信", "CH340 若 RTS 接 NRST：DTR=1、RTS=1，连接后等约 350ms"),
    ("九、线材", "41", "24V 电机线", "—", "电机供电", "0.75~1.5 mm²；尽量短，与信号线分开"),
    ("", "42", "STM32 信号线", "—", "GPIO/串口/SWD", "杜邦线；PA0~PA3、串口、SWD"),
    ("", "43", "继电器→切纸机", "—", "按钮并联线", "0.5 mm²；与电机线分开，可选屏蔽"),
    ("", "44", "共地", "—", "地线汇流", "较粗或星形汇流；避免地环路"),
]


def build_workbook() -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "接线总表"
    ws.merge_cells("A1:F1")
    ws["A1"].value = "CutPPaper 接线总表（nologo 0.96 TFT 一体板 · 四路继电器 · 115200 8N1 · 跳线 H）"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:F2")
    ws["A2"].value = "电机 H 桥：NO1–NO2→24V+，NC1–NC2→GND，COM1/COM2→伸缩杆 · 压纸：PA8 遮挡=压纸中"
    ws["A2"].font = Font(size=10, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center", wrap_text=True)

    headers = ["分类", "编号", "起点", "终点", "功能", "要求"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    style_header(ws, 3, 6)

    for i, (cat, num, src, dst, func, req) in enumerate(ROWS, 4):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=num)
        ws.cell(row=i, column=3, value=src)
        ws.cell(row=i, column=4, value=dst)
        ws.cell(row=i, column=5, value=func)
        ws.cell(row=i, column=6, value=req)
        style_row(ws, i, 6, alt=(i % 2 == 0))
        if cat:
            ws.cell(row=i, column=1).fill = cat_fill
            ws.cell(row=i, column=1).font = Font(bold=True)

    ws.freeze_panes = "A4"
    set_widths(ws, [16, 6, 28, 28, 18, 42])
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    ws2 = wb.create_sheet("电机逻辑")
    for c, h in enumerate(["PA0", "PA1", "结果", "COM1", "COM2"], 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, 5)
    for i, row in enumerate(
        [
            ("0", "0", "停止", "GND", "GND"),
            ("1", "0", "缩回", "+24V", "GND"),
            ("0", "1", "伸出", "GND", "+24V"),
            ("1", "1", "停止/刹车", "+24V", "+24V"),
        ],
        2,
    ):
        for c, v in enumerate(row, 1):
            ws2.cell(row=i, column=c, value=v)
        style_row(ws2, i, 5, alt=(i % 2 == 0))
    ws2.merge_cells("A7:E7")
    ws2["A7"] = "说明：任意 PA0/PA1 组合不会把 24V+ 与 GND 直接短接。"
    ws2["A7"].alignment = wrap
    set_widths(ws2, [8, 8, 14, 12, 12])

    ws3 = wb.create_sheet("软件步骤")
    for c, h in enumerate(["编号", "软件步骤", "STM32 命令", "硬件动作"], 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, 4)
    for i, row in enumerate(
        [
            ("A", "缩回", "RETRACT", "PA0=1 PA1=0，仅 K1；PA8 压纸中时自动停"),
            ("B", "继续", "PULSE_A", "PA2 脉冲，K3 吸合，COM3/NO3 闭合"),
            ("C", "切割", "PC 热键", "Cutting Master 4，无继电器"),
            ("D", "伸出", "EXTEND", "PA0=0 PA1=1，仅 K2"),
            ("E", "原点", "PULSE_B", "PA3 脉冲，K4 吸合，COM4/NO4 闭合"),
        ],
        2,
    ):
        for c, v in enumerate(row, 1):
            ws3.cell(row=i, column=c, value=v)
        style_row(ws3, i, 4, alt=(i % 2 == 0))
    set_widths(ws3, [8, 14, 14, 40])

    ws4 = wb.create_sheet("检查与禁止")
    ws4["A1"] = "上电前检查"
    ws4["A1"].font = Font(bold=True, size=12)
    checks = [
        "所有 GND 共地（24V−、STM32、TTL、ST-LINK、继电器 DC−）",
        "24V 不接 DC+；DC+ 接 12V（或 5V 模块接 5V）",
        "电机：NO1–NO2 接 24V+，NC1–NC2 接 GND，COM1/COM2 接电机；NC2 悬空",
        "串口 TX/RX 交叉，115200；VCC 不接 MCU",
        "跳线 S1~S4 插 H；K3/K4 并机器前确认干触点",
        "nologo 板：D3 串口 LED→PB9，切换键→PB8；PA6/PA7 为 LCD，勿外接",
        "槽型光电 DO→PA8；遮挡时 ROD_SENSOR 应返回 ROD:HOME（压纸中）",
        "先不接 24V 测 K1/K2；再空载试转；最后并机器按钮联调",
    ]
    for i, text in enumerate(checks, 2):
        ws4.cell(row=i, column=1, value=f"{i - 1}. {text}").alignment = wrap

    ws4["A10"] = "禁止事项"
    ws4["A10"].font = Font(bold=True, size=12, color="DC2626")
    for c, h in enumerate(["编号", "禁止", "后果"], 1):
        ws4.cell(row=11, column=c, value=h)
    style_header(ws4, 11, 3)
    for i, row in enumerate(
        [
            ("X1", "24V 接继电器 DC+", "烧线圈/光耦"),
            ("X2", "24V+ 接 STM32 引脚", "烧 MCU"),
            ("X3", "5V TTL 直连 STM32", "烧 MCU"),
            ("X4", "COM1=24V、COM2=GND 旧接法", "停止时电机可能仍通电"),
            ("X5", "NC 与 NO 跨继电器并到同一电机线", "易短路"),
            ("X6", "强电按钮直接并 COM/NO", "须干触点或光耦隔离"),
        ],
        12,
    ):
        for c, v in enumerate(row, 1):
            ws4.cell(row=i, column=c, value=v)
        style_row(ws4, i, 3, alt=(i % 2 == 0))
    set_widths(ws4, [55, 32, 24])

    return wb


if __name__ == "__main__":
    build_workbook().save(OUT)
    print(f"已生成: {OUT}")
